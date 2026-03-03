import os
import time
import random
from datetime import datetime
from io import BytesIO

import pandas as pd
import streamlit as st

# PDF + gráficos
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import A4

import matplotlib.pyplot as plt

# Excel PRO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CSV_PATH = "sensores.csv"

# -------------------- Rangos OK / ALERT --------------------
RANGES_OK = {
    "rpm": (650, 3500),
    "speed": (0, 140),
    "coolant_C": (75, 105),
    "voltage_V": (12.0, 14.8),
    "map_kPa": (20, 220),
    "lambda": (0.90, 1.10),
    "maf_gps": (1.0, 80.0),
    "fuel_trim_pct": (-10.0, 10.0),
}

RANGES_ALERT = {
    "coolant_C": (70, 110),
    "voltage_V": (11.8, 14.9),
    "lambda": (0.85, 1.20),
    "fuel_trim_pct": (-20.0, 20.0),
}

SENSOR_META = {
    "rpm": ("RPM", "rpm"),
    "speed": ("Velocidad", "km/h"),
    "coolant_C": ("Temp. refrigerante", "°C"),
    "voltage_V": ("Voltaje batería", "V"),
    "map_kPa": ("MAP (presión admisión)", "kPa"),
    "lambda": ("Lambda", ""),
    "maf_gps": ("MAF (flujo de aire)", "g/s"),
    "fuel_trim_pct": ("Fuel Trim", "%"),
}

# -------------------- Diccionario DTC (explicación técnica) --------------------
DTC_INFO = {
    "P0101": "MAF rango/rendimiento. Posible MAF sucio, fuga admisión, filtro, cableado, incoherencia con MAP.",
    "P0106": "MAP rango/rendimiento. Posible sensor MAP, mangueras, actuador, cableado o presión incoherente.",
    "P0130": "Circuito sonda lambda (O2). Posible sonda, cableado, fugas escape, mezcla fuera de control.",
    "P0171": "Mezcla pobre (Bank 1). Posible fuga aire, baja presión combustible, MAF subestima, inyectores sucios.",
    "P0172": "Mezcla rica (Bank 1). Posible presión alta, inyectores goteando, EVAP, MAF sobreestima.",
    "P0217": "Temperatura motor excesiva. Posible refrigeración: nivel, termostato, ventilador, bomba.",
    "P0300": "Fallo de encendido aleatorio. Posible bujías/bobinas, inyección, compresión, fugas.",
    "P0420": "Eficiencia catalizador baja. Posible catalizador, sonda O2, mezcla rica/pobre prolongada.",
    "P0562": "Tensión sistema baja. Posible batería, alternador, masa/borne, consumo parasitario.",
}

# -------------------- Modos de prueba --------------------
MODE_PROFILES = {
    "Ralentí": {"speed_target": 0, "rpm_base": 820, "rpm_var": 80, "map_base": 35},
    "Ciudad": {"speed_target": 35, "rpm_base": 1400, "rpm_var": 250, "map_base": 55},
    "Autopista": {"speed_target": 100, "rpm_base": 2400, "rpm_var": 200, "map_base": 95},
    "Aceleración": {"speed_target": 80, "rpm_base": 2800, "rpm_var": 400, "map_base": 130},
    "Frenada": {"speed_target": 20, "rpm_base": 1200, "rpm_var": 200, "map_base": 45},
}

FAULTS = [
    "Ninguno",
    "Batería baja (voltaje)",
    "Sonda lambda fuera de rango",
    "Sobrecalentamiento",
    "MAP anómalo",
    "Fuel trim alto (mezcla pobre/rica)",
    "DTC intermitente",
]

# -------------------- Utilidades --------------------
def clamp(x, lo, hi):
    return max(lo, min(hi, x))

def now_str():
    return datetime.now().isoformat(timespec="seconds")

def sensor_state(value, key):
    ok_lo, ok_hi = RANGES_OK[key]
    al_lo, al_hi = RANGES_ALERT.get(key, (ok_lo, ok_hi))
    if value < al_lo or value > al_hi:
        return "ALERT"
    if value < ok_lo or value > ok_hi:
        return "WARN"
    return "OK"

EXPECTED_COLS = ["time","mode","rpm","speed","coolant_C","voltage_V","map_kPa","lambda","maf_gps","fuel_trim_pct","dtc"]

def load_df():
    if not os.path.exists(CSV_PATH):
        return pd.DataFrame(columns=EXPECTED_COLS)

    df = pd.read_csv(CSV_PATH)

    # compatibilidad con CSV antiguo
    for col in EXPECTED_COLS:
        if col not in df.columns:
            if col == "dtc":
                df[col] = "OK"
            elif col == "mode":
                df[col] = "Ralentí"
            else:
                df[col] = pd.NA

    return df[EXPECTED_COLS]

def save_row(row):
    df = load_df()
    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    df.to_csv(CSV_PATH, index=False)

def add_event(msg, level="INFO"):
    st.session_state.events.append({"time": now_str(), "level": level, "msg": msg})

def parse_dtc_code(dtc_str: str) -> str:
    if not dtc_str or dtc_str == "OK":
        return ""
    # "P0101 (....)" -> "P0101"
    token = dtc_str.strip().split()[0]
    if token.startswith("P") and len(token) >= 5:
        return token[:5]
    # "P0171/P0172 ..." -> devuelve el primero
    if "/" in token and token.split("/")[0].startswith("P"):
        return token.split("/")[0][:5]
    return ""

# -------------------- Simulación --------------------
def apply_fault(row, fault):
    if fault == "Batería baja (voltaje)":
        row["voltage_V"] = round(random.uniform(11.2, 11.7), 2)
        row["dtc"] = "P0562 (Tensión sistema baja)"
    elif fault == "Sonda lambda fuera de rango":
        row["lambda"] = round(random.uniform(1.18, 1.28), 2)
        row["dtc"] = "P0130 (Sonda lambda circuito)"
    elif fault == "Sobrecalentamiento":
        row["coolant_C"] = round(random.uniform(109, 114), 1)
        row["dtc"] = "P0217 (Temperatura motor excesiva)"
    elif fault == "MAP anómalo":
        row["map_kPa"] = int(random.uniform(160, 210))
        row["dtc"] = "P0106 (MAP rango/rendimiento)"
    elif fault == "Fuel trim alto (mezcla pobre/rica)":
        row["fuel_trim_pct"] = round(random.choice([random.uniform(15, 25), random.uniform(-25, -15)]), 1)
        row["dtc"] = "P0171/P0172 (Mezcla pobre/rica)"
    elif fault == "DTC intermitente":
        if random.random() < 0.7:
            row["dtc"] = random.choice([
                "P0101 (MAF/MAP incoherente)",
                "P0420 (Eficiencia catalizador baja)",
                "P0300 (Fallo encendido aleatorio)"
            ])
    return row

def generate_sample(prev, mode, forced_fault):
    prof = MODE_PROFILES[mode]

    if prev is None:
        speed = 0
        rpm = int(random.gauss(prof["rpm_base"], prof["rpm_var"]))
        coolant = round(random.gauss(88, 2), 1)
    else:
        target = prof["speed_target"]
        speed = int(clamp(prev["speed"] + (target - prev["speed"]) * random.uniform(0.15, 0.30) + random.gauss(0, 3), 0, 140))
        rpm = int(clamp(random.gauss(prof["rpm_base"] + speed * 6, prof["rpm_var"]), 650, 3800))
        coolant = round(clamp(prev["coolant_C"] + random.gauss(0.05, 0.12), 70, 112), 1)

    voltage = round(clamp(random.gauss(13.9, 0.25), 12.2, 14.8), 2)
    map_kpa = int(clamp(random.gauss(prof["map_base"] + speed * 0.6, 10), 20, 220))
    maf_gps = round(clamp(random.gauss(2.0 + rpm * 0.015 + (map_kpa/50), 2.0), 0.5, 90.0), 2)
    lambda_eq = round(clamp(random.gauss(1.00, 0.03), 0.85, 1.25), 2)
    fuel_trim = round(clamp(random.gauss(0.0, 4.0), -25.0, 25.0), 1)

    dtc = "OK"
    if random.random() < 0.03:
        dtc = random.choice(["P0420 (Eficiencia catalizador baja)", "P0101 (MAF/MAP incoherente)"])

    row = {
        "time": now_str(),
        "mode": mode,
        "rpm": rpm,
        "speed": speed,
        "coolant_C": coolant,
        "voltage_V": voltage,
        "map_kPa": map_kpa,
        "lambda": lambda_eq,
        "maf_gps": maf_gps,
        "fuel_trim_pct": fuel_trim,
        "dtc": dtc
    }

    if forced_fault != "Ninguno":
        row = apply_fault(row, forced_fault)

    return row

# -------------------- Coherencias (diagnosis inteligente) --------------------
def coherence_checks(row):
    issues = []
    rpm = row["rpm"]
    speed = row["speed"]
    map_kpa = row["map_kPa"]
    maf = row["maf_gps"]
    voltage = row["voltage_V"]
    coolant = row["coolant_C"]
    lam = row["lambda"]
    trim = row["fuel_trim_pct"]

    if speed > 10 and rpm < 700:
        issues.append(("ALERT", "Velocidad > 10 km/h con RPM muy bajas (posible fallo sensor RPM)."))
    if map_kpa > 160 and rpm < 1400:
        issues.append(("WARN", "MAP alto con RPM bajas (posible fallo MAP/actuador o fuga/actuador)."))
    if maf < 2.0 and rpm > 2000:
        issues.append(("WARN", "MAF bajo con RPM altas (posible MAF sucio o restricción admisión)."))
    if voltage < 12.0:
        issues.append(("ALERT", "Voltaje bajo: revisar batería/alternador."))
    if coolant > 108:
        issues.append(("ALERT", "Temperatura refrigerante alta: revisar refrigeración."))
    if lam > 1.15 or lam < 0.85:
        issues.append(("ALERT", "Lambda fuera de rango: revisar sonda/mezcla."))
    if abs(trim) > 15:
        issues.append(("WARN", "Fuel Trim elevado: posible fuga/toma de aire/inyectores/MAF."))

    return issues

# -------------------- Diagnóstico automático (explicación técnica) --------------------
def auto_diagnosis(last: dict, issues: list) -> list:
    """
    Devuelve lista de dicts: {Categoria, Hallazgo, Explicacion, Severidad}
    """
    out = []

    dtc = str(last.get("dtc", "OK"))
    code = parse_dtc_code(dtc)
    if dtc != "OK":
        exp = DTC_INFO.get(code, "Código DTC detectado. Consultar tabla del fabricante y seguir procedimiento de diagnosis.")
        out.append({
            "Categoria": "DTC",
            "Hallazgo": dtc,
            "Explicacion": exp,
            "Severidad": "WARN" if "intermit" in dtc.lower() else "ALERT"
        })

    # Fuel trim (mezcla)
    ft = float(last.get("fuel_trim_pct", 0))
    if abs(ft) >= 15:
        if ft > 0:
            out.append({
                "Categoria": "Mezcla (Fuel Trim)",
                "Hallazgo": f"Fuel Trim +{ft:.1f}%",
                "Explicacion": "Mezcla pobre: la ECU añade combustible. Causas típicas: fuga de aire en admisión, MAF subestimando, baja presión de combustible, inyectores obstruidos.",
                "Severidad": "WARN"
            })
        else:
            out.append({
                "Categoria": "Mezcla (Fuel Trim)",
                "Hallazgo": f"Fuel Trim {ft:.1f}%",
                "Explicacion": "Mezcla rica: la ECU reduce combustible. Causas típicas: inyectores goteando, presión combustible alta, EVAP, MAF sobreestimando.",
                "Severidad": "WARN"
            })

    # Lambda fuera de rango
    lam = float(last.get("lambda", 1.0))
    if lam > 1.15:
        out.append({
            "Categoria": "Lambda",
            "Hallazgo": f"Lambda alta ({lam:.2f})",
            "Explicacion": "Indica tendencia a mezcla pobre o lectura anómala. Revisar fugas de admisión/escape y estado de sonda lambda.",
            "Severidad": "ALERT"
        })
    elif lam < 0.85:
        out.append({
            "Categoria": "Lambda",
            "Hallazgo": f"Lambda baja ({lam:.2f})",
            "Explicacion": "Indica mezcla rica o lectura anómala. Revisar presión de combustible, inyectores, EVAP y sonda lambda.",
            "Severidad": "ALERT"
        })

    # Incoherencias MAF/MAP
    mapk = int(last.get("map_kPa", 0))
    rpm = int(last.get("rpm", 0))
    maf = float(last.get("maf_gps", 0))
    if mapk > 160 and rpm < 1400:
        out.append({
            "Categoria": "Coherencia MAF/MAP",
            "Hallazgo": f"MAP alto ({mapk} kPa) con RPM bajas ({rpm})",
            "Explicacion": "Posible sensor MAP defectuoso, actuador/boost incoherente o fuga. Revisar sensor/conexiones y mangueras.",
            "Severidad": "WARN"
        })
    if maf < 2.0 and rpm > 2000:
        out.append({
            "Categoria": "Coherencia MAF/MAP",
            "Hallazgo": f"MAF bajo ({maf:.2f} g/s) con RPM altas ({rpm})",
            "Explicacion": "Posible MAF sucio, filtro de aire obstruido o restricción en admisión. Revisar filtro y limpiar/probar MAF.",
            "Severidad": "WARN"
        })

    # Reflejar issues de coherencia (lo que ya calculabas)
    for level, msg in issues:
        out.append({
            "Categoria": "Coherencias",
            "Hallazgo": msg,
            "Explicacion": "Regla automática: patrón de sensores incoherente para ese modo de funcionamiento.",
            "Severidad": level
        })

    # Si no hay nada, estado OK
    if not out:
        out.append({
            "Categoria": "OK",
            "Hallazgo": "Sin incidencias relevantes",
            "Explicacion": "Valores dentro de rango y coherencias correctas para el modo seleccionado.",
            "Severidad": "OK"
        })

    return out

# -------------------- Score 0-100 --------------------
def compute_health_score(df):
    score = 100
    warn = alert = dtc_count = coh_warn = coh_alert = 0

    if len(df) == 0:
        return 0, 0, 0, 0, 0, 0

    for _, r in df.iterrows():
        for key in SENSOR_META.keys():
            val = float(r[key]) if key in ["coolant_C","voltage_V","lambda","maf_gps","fuel_trim_pct"] else int(r[key])
            stt = sensor_state(val, key)
            if stt == "WARN":
                warn += 1
            elif stt == "ALERT":
                alert += 1

        if str(r["dtc"]) != "OK":
            dtc_count += 1

        issues = coherence_checks({
            "rpm": int(r["rpm"]),
            "speed": int(r["speed"]),
            "map_kPa": int(r["map_kPa"]),
            "maf_gps": float(r["maf_gps"]),
            "voltage_V": float(r["voltage_V"]),
            "coolant_C": float(r["coolant_C"]),
            "lambda": float(r["lambda"]),
            "fuel_trim_pct": float(r["fuel_trim_pct"]),
        })
        for level, _ in issues:
            if level == "WARN":
                coh_warn += 1
            else:
                coh_alert += 1

    score -= warn * 0.3
    score -= alert * 1.2
    score -= dtc_count * 2.0
    score -= coh_warn * 0.8
    score -= coh_alert * 2.5
    score = int(clamp(score, 0, 100))
    return score, warn, alert, dtc_count, coh_warn, coh_alert

def score_band(score):
    if score >= 85:
        return "EXCELENTE", colors.HexColor("#2E7D32")
    if score >= 70:
        return "ACEPTABLE", colors.HexColor("#F9A825")
    if score >= 50:
        return "RIESGO MEDIO", colors.HexColor("#EF6C00")
    return "RIESGO ALTO", colors.HexColor("#C62828")

# -------------------- Gráfico para PDF --------------------
def make_pdf_chart(df_last60: pd.DataFrame) -> BytesIO:
    fig = plt.figure(figsize=(6.8, 2.3), dpi=150)
    ax = fig.add_subplot(111)

    x = list(range(len(df_last60)))
    ax.plot(x, df_last60["rpm"].astype(float), label="RPM")
    ax.plot(x, df_last60["coolant_C"].astype(float), label="Temp (°C)")
    ax.plot(x, df_last60["voltage_V"].astype(float), label="Voltaje (V)")

    ax.set_title("Tendencia (últimos 60 registros)", fontsize=9)
    ax.set_xlabel("Muestras", fontsize=8)
    ax.grid(True, alpha=0.25)
    ax.legend(fontsize=7, ncols=3, loc="upper left")

    buf = BytesIO()
    fig.tight_layout()
    fig.savefig(buf, format="png")
    plt.close(fig)
    buf.seek(0)
    return buf

# -------------------- PDF PRO mejorado + diagnóstico automático --------------------
def build_pdf_report_pro(df, last, score, warn_total, alert_total, dtc_total, recs, vehicle, events_df, diag_rows) -> bytes:
    buf = BytesIO()

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillGray(0.35)
        canvas.drawString(36, 20, f"AutoCheck · Informe generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        canvas.drawRightString(A4[0]-36, 20, f"Página {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=42
    )

    styles = getSampleStyleSheet()
    H2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=12, spaceBefore=10, spaceAfter=6)
    normal = ParagraphStyle("normal", parent=styles["Normal"], fontSize=9.8, leading=13)

    band_text, band_color = score_band(score)

    started = vehicle.get("started_at") or (df.iloc[0]["time"] if len(df) else now_str())
    ended = df.iloc[-1]["time"] if len(df) else now_str()

    elements = []

    # Cabecera
    header = Table(
        [["AutoCheck", "INFORME DE DIAGNOSIS ELECTRÓNICA"]],
        colWidths=[2.3*inch, 3.7*inch]
    )
    header.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0,0), (-1,-1), colors.white),
        ("FONTNAME", (0,0), (0,0), "Helvetica-Bold"),
        ("FONTNAME", (1,0), (1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (0,0), 16),
        ("FONTSIZE", (1,0), (1,0), 11),
        ("ALIGN", (0,0), (0,0), "LEFT"),
        ("ALIGN", (1,0), (1,0), "RIGHT"),
        ("LEFTPADDING", (0,0), (-1,-1), 10),
        ("RIGHTPADDING", (0,0), (-1,-1), 10),
        ("TOPPADDING", (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
    ]))
    elements.append(header)
    elements.append(Spacer(1, 0.15*inch))
    elements.append(Paragraph("Monitor Multisensor (Simulación) · Lectura simultánea · Diagnosis automática · Reporte", normal))
    elements.append(Spacer(1, 0.15*inch))

    # Caja score
    score_box = Table(
        [[
            Paragraph(f"<b>Índice de salud:</b> {score}/100", normal),
            Paragraph(f"<b>Clasificación:</b> {band_text}", normal),
            Paragraph(f"<b>WARN:</b> {warn_total}  <b>ALERT:</b> {alert_total}  <b>DTC:</b> {dtc_total}", normal),
        ]],
        colWidths=[2.2*inch, 1.7*inch, 2.1*inch]
    )
    score_box.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.whitesmoke),
        ("BOX", (0,0), (-1,-1), 1.2, band_color),
        ("LINEBEFORE", (1,0), (1,0), 1, colors.lightgrey),
        ("LINEBEFORE", (2,0), (2,0), 1, colors.lightgrey),
        ("LEFTPADDING", (0,0), (-1,-1), 10),
        ("RIGHTPADDING", (0,0), (-1,-1), 10),
        ("TOPPADDING", (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
    ]))
    elements.append(score_box)
    elements.append(Spacer(1, 0.18*inch))

    # Datos generales
    elements.append(Paragraph("1. Datos generales", H2))
    general = [
        ["Fecha inicio", str(started)],
        ["Fecha fin", str(ended)],
        ["Operador", vehicle.get("operador","") or "-"],
        ["Vehículo", vehicle.get("marca_modelo","") or "-"],
        ["Motor", vehicle.get("motor","") or "-"],
        ["Kilometraje", f"{vehicle.get('km',0)} km"],
        ["Matrícula", vehicle.get("matricula","") or "-"],
        ["VIN", vehicle.get("vin","") or "-"],
    ]
    t_general = Table(general, colWidths=[1.7*inch, 4.3*inch])
    t_general.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.4, colors.lightgrey),
        ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#F3F4F6")),
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 9.8),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    elements.append(t_general)

    # Últimos valores
    elements.append(Paragraph("2. Últimos valores registrados", H2))
    values = [["Parámetro", "Valor"]]
    for key, (label, unit) in SENSOR_META.items():
        v = float(last[key]) if key in ["coolant_C","voltage_V","lambda","maf_gps","fuel_trim_pct"] else int(last[key])
        unit_txt = f" {unit}" if unit else ""
        if key == "lambda":
            txt = f"{v:.2f}{unit_txt}"
        elif key == "coolant_C":
            txt = f"{v:.1f}{unit_txt}"
        elif key in ["voltage_V","maf_gps"]:
            txt = f"{v:.2f}{unit_txt}"
        elif key == "fuel_trim_pct":
            txt = f"{v:.1f}{unit_txt}"
        else:
            txt = f"{v}{unit_txt}"
        values.append([label, txt])
    values.append(["Modo", str(last.get("mode","-"))])
    values.append(["DTC", str(last.get("dtc","OK"))])

    t_values = Table(values, colWidths=[2.7*inch, 3.3*inch])
    t_values.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.4, colors.lightgrey),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("FONTSIZE", (0,0), (-1,-1), 9.8),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    elements.append(t_values)

    # Estado por sensor (colores)
    elements.append(Paragraph("3. Estado por sensor", H2))
    state_rows = [["Sensor", "Valor", "Estado"]]
    for key, (label, unit) in SENSOR_META.items():
        v = float(last[key]) if key in ["coolant_C","voltage_V","lambda","maf_gps","fuel_trim_pct"] else int(last[key])
        stt = sensor_state(v, key)
        unit_txt = f" {unit}" if unit else ""
        if key == "coolant_C":
            v_txt = f"{v:.1f}{unit_txt}"
        elif key in ["voltage_V","maf_gps"]:
            v_txt = f"{v:.2f}{unit_txt}"
        elif key == "lambda":
            v_txt = f"{v:.2f}"
        elif key == "fuel_trim_pct":
            v_txt = f"{v:.1f}{unit_txt}"
        else:
            v_txt = f"{v}{unit_txt}"
        state_rows.append([label, v_txt, stt])

    t_state = Table(state_rows, colWidths=[2.7*inch, 2.0*inch, 1.3*inch])
    ts = TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.4, colors.lightgrey),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("FONTSIZE", (0,0), (-1,-1), 9.8),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ])
    for i in range(1, len(state_rows)):
        state = state_rows[i][2]
        if state == "OK":
            bg, fg = colors.HexColor("#E8F5E9"), colors.HexColor("#1B5E20")
        elif state == "WARN":
            bg, fg = colors.HexColor("#FFF8E1"), colors.HexColor("#F57F17")
        else:
            bg, fg = colors.HexColor("#FFEBEE"), colors.HexColor("#B71C1C")
        ts.add("BACKGROUND", (2,i), (2,i), bg)
        ts.add("TEXTCOLOR", (2,i), (2,i), fg)
        ts.add("FONTNAME", (2,i), (2,i), "Helvetica-Bold")
    t_state.setStyle(ts)
    elements.append(t_state)

    # Diagnóstico automático (NUEVO)
    elements.append(Paragraph("4. Diagnóstico automático (explicación técnica)", H2))
    diag_table_rows = [["Categoría", "Hallazgo", "Explicación", "Sev."]]
    for r in diag_rows:
        diag_table_rows.append([
            str(r.get("Categoria","")),
            str(r.get("Hallazgo","")),
            str(r.get("Explicacion","")),
            str(r.get("Severidad","")),
        ])

    t_diag = Table(diag_table_rows, colWidths=[1.0*inch, 1.7*inch, 2.9*inch, 0.5*inch])
    td = TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.35, colors.lightgrey),
        ("FONTSIZE", (0,0), (-1,-1), 8.6),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("LEFTPADDING", (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ])
    # colorear columna severidad
    for i in range(1, len(diag_table_rows)):
        sev = diag_table_rows[i][3]
        if sev == "OK":
            td.add("BACKGROUND", (3,i), (3,i), colors.HexColor("#E8F5E9"))
            td.add("TEXTCOLOR", (3,i), (3,i), colors.HexColor("#1B5E20"))
        elif sev == "WARN":
            td.add("BACKGROUND", (3,i), (3,i), colors.HexColor("#FFF8E1"))
            td.add("TEXTCOLOR", (3,i), (3,i), colors.HexColor("#F57F17"))
        else:
            td.add("BACKGROUND", (3,i), (3,i), colors.HexColor("#FFEBEE"))
            td.add("TEXTCOLOR", (3,i), (3,i), colors.HexColor("#B71C1C"))
        td.add("FONTNAME", (3,i), (3,i), "Helvetica-Bold")
    t_diag.setStyle(td)
    elements.append(t_diag)

    # Gráfico tendencia
    elements.append(Paragraph("5. Gráfico de tendencia", H2))
    chart_buf = make_pdf_chart(df.tail(60))
    img = RLImage(chart_buf, width=6.1*inch, height=2.0*inch)
    img.hAlign = "CENTER"
    elements.append(img)

    # Recomendaciones
    elements.append(Paragraph("6. Recomendaciones técnicas", H2))
    for r in recs:
        elements.append(Paragraph(f"• {r}", normal))

    # Eventos
    elements.append(Paragraph("7. Registro de eventos", H2))
    if events_df is not None and len(events_df) > 0:
        tail = events_df.tail(10).copy()
        ev_rows = [["Hora", "Nivel", "Evento"]]
        for _, rr in tail.iterrows():
            ev_rows.append([str(rr.get("time","")), str(rr.get("level","")), str(rr.get("msg",""))])

        t_ev = Table(ev_rows, colWidths=[1.5*inch, 0.9*inch, 3.6*inch])
        t_ev.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#111827")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("GRID", (0,0), (-1,-1), 0.4, colors.lightgrey),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F9FAFB")]),
            ("FONTSIZE", (0,0), (-1,-1), 9),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("TOPPADDING", (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ]))
        elements.append(t_ev)
    else:
        elements.append(Paragraph("No se registraron eventos relevantes.", normal))

    # Veredicto final
    elements.append(Paragraph("8. Veredicto final", H2))
    if score >= 85 and alert_total == 0 and dtc_total == 0:
        verdict = "Estado electrónico muy bueno. No se observan incidencias relevantes."
    elif score >= 70:
        verdict = "Estado aceptable. Se recomienda revisión preventiva y seguimiento de incidencias."
    elif score >= 50:
        verdict = "Riesgo medio. Se recomienda diagnosis ampliada antes de entregar o vender el vehículo."
    else:
        verdict = "Riesgo alto. Se recomienda revisión urgente y no circular hasta verificar fallos."

    verdict_box = Table([[Paragraph(f"<b>{verdict}</b>", normal)]], colWidths=[6.0*inch])
    verdict_box.setStyle(TableStyle([
        ("BOX", (0,0), (-1,-1), 1.2, band_color),
        ("BACKGROUND", (0,0), (-1,-1), colors.whitesmoke),
        ("LEFTPADDING", (0,0), (-1,-1), 10),
        ("TOPPADDING", (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
    ]))
    elements.append(verdict_box)

    doc.build(elements, onFirstPage=footer, onLaterPages=footer)
    return buf.getvalue()

# -------------------- Excel PRO (xlsx) --------------------
def make_excel_pro(df: pd.DataFrame, vehicle: dict, score: int, warn_total: int, alert_total: int, dtc_total: int, diag_rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sesion"

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    ok_fill = PatternFill("solid", fgColor="E8F5E9")
    warn_fill = PatternFill("solid", fgColor="FFF8E1")
    alert_fill = PatternFill("solid", fgColor="FFEBEE")

    # Orden columnas
    order = ["time","mode","rpm","speed","coolant_C","voltage_V","map_kPa","lambda","maf_gps","fuel_trim_pct","dtc"]
    dfx = df.copy()
    for c in order:
        if c not in dfx.columns:
            dfx[c] = ""
    dfx = dfx[order]

    # Escribir cabecera
    ws.append(order)
    for j, col in enumerate(order, start=1):
        cell = ws.cell(row=1, column=j)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Filas
    for _, r in dfx.iterrows():
        ws.append([r[c] for c in order])

    ws.freeze_panes = "A2"

    # Ajustar anchos
    for col_idx, col_name in enumerate(order, start=1):
        max_len = len(col_name)
        for i in range(2, min(ws.max_row, 250) + 1):
            v = ws.cell(row=i, column=col_idx).value
            max_len = max(max_len, len(str(v)) if v is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)

    # Colorear DTC
    dtc_col = order.index("dtc") + 1
    for i in range(2, ws.max_row + 1):
        dtc_val = str(ws.cell(row=i, column=dtc_col).value)
        if dtc_val and dtc_val != "OK":
            ws.cell(row=i, column=dtc_col).fill = alert_fill
            ws.cell(row=i, column=dtc_col).font = Font(bold=True, color="B71C1C")

    # Hoja Resumen
    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "AutoCheck — Resumen de diagnóstico"
    ws2["A1"].font = Font(bold=True, size=16)
    ws2["A3"] = "Índice de salud"
    ws2["B3"] = f"{score}/100"
    ws2["A4"] = "WARN"
    ws2["B4"] = warn_total
    ws2["A5"] = "ALERT"
    ws2["B5"] = alert_total
    ws2["A6"] = "DTC detectados"
    ws2["B6"] = dtc_total

    # Datos vehículo
    ws2["D3"] = "Operador"
    ws2["E3"] = vehicle.get("operador","") or "-"
    ws2["D4"] = "Vehículo"
    ws2["E4"] = vehicle.get("marca_modelo","") or "-"
    ws2["D5"] = "Motor"
    ws2["E5"] = vehicle.get("motor","") or "-"
    ws2["D6"] = "KM"
    ws2["E6"] = vehicle.get("km",0)

    # Diagnóstico automático
    ws2["A8"] = "Diagnóstico automático (explicación técnica)"
    ws2["A8"].font = Font(bold=True, size=12)

    diag_headers = ["Categoría", "Hallazgo", "Explicación", "Sev."]
    ws2.append([])  # fila vacía
    ws2.append(diag_headers)
    header_row = ws2.max_row
    for j in range(1, 5):
        c = ws2.cell(row=header_row, column=j)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    for r in diag_rows:
        ws2.append([r.get("Categoria",""), r.get("Hallazgo",""), r.get("Explicacion",""), r.get("Severidad","")])

    # Formato diagnóstico
    for i in range(header_row + 1, ws2.max_row + 1):
        ws2.cell(row=i, column=1).alignment = wrap
        ws2.cell(row=i, column=2).alignment = wrap
        ws2.cell(row=i, column=3).alignment = wrap
        sev = str(ws2.cell(row=i, column=4).value)
        ws2.cell(row=i, column=4).alignment = center
        if sev == "OK":
            ws2.cell(row=i, column=4).fill = ok_fill
        elif sev == "WARN":
            ws2.cell(row=i, column=4).fill = warn_fill
        else:
            ws2.cell(row=i, column=4).fill = alert_fill
            ws2.cell(row=i, column=4).font = Font(bold=True, color="B71C1C")

    # Ajustar anchos resumen
    widths = {1: 18, 2: 28, 3: 65, 4: 8, 5: 24}
    for col, w in widths.items():
        ws2.column_dimensions[get_column_letter(col)].width = w

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

# -------------------- STREAMLIT UI --------------------
st.set_page_config(page_title="AutoCheck - Multisensor", layout="wide")
st.title("AutoCheck — Monitor Multisensor (SIMULACIÓN)")
st.caption("Lectura simultánea + grabación + diagnosis + score 0–100 + informe PDF profesional + Excel PRO.")

if "running" not in st.session_state:
    st.session_state.running = False
if "last" not in st.session_state:
    st.session_state.last = None
if "events" not in st.session_state:
    st.session_state.events = []
if "started_at" not in st.session_state:
    st.session_state.started_at = None

with st.expander("📋 Datos del vehículo (para el PDF)", expanded=False):
    c1, c2, c3 = st.columns(3)
    with c1:
        st.session_state.vin = st.text_input("VIN (opcional)", value=st.session_state.get("vin",""))
        st.session_state.matricula = st.text_input("Matrícula (opcional)", value=st.session_state.get("matricula",""))
    with c2:
        st.session_state.marca_modelo = st.text_input("Marca/Modelo", value=st.session_state.get("marca_modelo",""))
        st.session_state.km = st.number_input("Kilometraje (km)", min_value=0, value=int(st.session_state.get("km",0)))
    with c3:
        st.session_state.motor = st.selectbox("Tipo motor", ["Gasolina","Diésel","Híbrido","Eléctrico"], index=0)
        st.session_state.operador = st.text_input("Operador (tu nombre)", value=st.session_state.get("operador",""))

top1, top2, top3, top4, top5 = st.columns([1.2, 1.2, 1.0, 1.2, 1.2])
with top1:
    mode = st.selectbox("Modo de prueba", list(MODE_PROFILES.keys()))
with top2:
    forced_fault = st.selectbox("Inyección de fallo (demo)", FAULTS, index=0)
with top3:
    auto = st.toggle("Auto-actualización (1s)", value=False)
with top4:
    if st.button("▶ Iniciar sesión", use_container_width=True):
        st.session_state.running = True
        if st.session_state.started_at is None:
            st.session_state.started_at = now_str()
        add_event(f"Sesión iniciada. Modo: {mode}.", "INFO")
with top5:
    if st.button("⏸ Parar sesión", use_container_width=True):
        st.session_state.running = False
        add_event("Sesión detenida.", "INFO")

c_clear, c_tip = st.columns([1, 2])
with c_clear:
    if st.button("🧹 Borrar sesión (sensores.csv)", use_container_width=True):
        if os.path.exists(CSV_PATH):
            os.remove(CSV_PATH)
        st.session_state.last = None
        st.session_state.events = []
        st.session_state.started_at = None
        st.success("Sesión borrada.")
with c_tip:
    st.caption("Para demostrar la innovación: activa un fallo → verás WARN/ALERT + baja el score y queda reflejado en el PDF y Excel PRO.")

df = load_df()

# generar un registro por refresco si está corriendo
if st.session_state.running:
    prev = st.session_state.last if isinstance(st.session_state.last, dict) else None
    new_row = generate_sample(prev, mode, forced_fault)
    st.session_state.last = new_row
    save_row(new_row)
    df = load_df()

if len(df) == 0:
    st.info("Pulsa **Iniciar sesión** para empezar a generar datos.")
else:
    last = df.iloc[-1].to_dict()

    score, warn_n, alert_n, dtc_n, coh_warn, coh_alert = compute_health_score(df)
    warn_total = int(warn_n + coh_warn)
    alert_total = int(alert_n + coh_alert)

    # Issues del último instante
    issues = coherence_checks({
        "rpm": int(last["rpm"]),
        "speed": int(last["speed"]),
        "map_kPa": int(last["map_kPa"]),
        "maf_gps": float(last["maf_gps"]),
        "voltage_V": float(last["voltage_V"]),
        "coolant_C": float(last["coolant_C"]),
        "lambda": float(last["lambda"]),
        "fuel_trim_pct": float(last["fuel_trim_pct"]),
    })

    # Log de eventos
    if str(last["dtc"]) != "OK":
        add_event(f"DTC detectado: {last['dtc']}", "WARN")
    for level, msg in issues:
        add_event(msg, level)

    # Diagnóstico automático (nuevo)
    diag_rows = auto_diagnosis(last, issues)
    diag_df = pd.DataFrame(diag_rows)

    # Recomendaciones (mejoradas con diagnóstico)
    recs = []
    if dtc_n > 0:
        recs.append("DTC presentes: borrar/leer, comprobar si reaparecen y seguir procedimiento del fabricante.")
    if alert_total > 0:
        recs.append("Valores críticos: revisar sistema eléctrico/refrigeración/mezcla según el sensor afectado.")
    if any(r.get("Categoria","").startswith("Mezcla") for r in diag_rows):
        recs.append("Si hay mezcla pobre/rica: revisar fugas admisión, presión combustible, MAF y estado de inyectores.")
    if any("MAF" in str(r.get("Hallazgo","")) or "MAP" in str(r.get("Hallazgo","")) for r in diag_rows):
        recs.append("Si hay incoherencias MAF/MAP: revisar admisión (filtro/fugas), sensores y conectores.")
    if score < 70:
        recs.append("Ampliar diagnosis: prueba dinámica, comprobación de fugas y verificación MAP/MAF/Lambda.")
    if not recs:
        recs.append("Sin incidencias destacables: mantenimiento preventivo y seguimiento periódico.")

    # MÉTRICAS
    band_text, _ = score_band(score)
    a, b, c, d, e = st.columns([1.4, 1, 1.6, 0.8, 0.8])
    with a:
        st.metric("Índice de salud electrónica", f"{score}/100", band_text)
        st.progress(score)
    with b:
        st.metric("Modo", str(last["mode"]))
    with c:
        st.metric("DTC", str(last["dtc"]))
    with d:
        st.metric("WARN", warn_total)
    with e:
        st.metric("ALERT", alert_total)

    st.divider()

    # Estado por sensor
    st.subheader("Estado por sensor (OK / WARN / ALERT)")
    rows = []
    for key, (label, unit) in SENSOR_META.items():
        v = float(last[key]) if key in ["coolant_C","voltage_V","lambda","maf_gps","fuel_trim_pct"] else int(last[key])
        rows.append([label, v, sensor_state(v, key), unit])
    st.dataframe(pd.DataFrame(rows, columns=["Sensor", "Valor", "Estado", "Unidad"]), use_container_width=True)

    # Diagnosis coherencias
    st.subheader("Diagnosis inteligente (coherencias)")
    if issues:
        st.warning("Se han detectado incoherencias. Quedarán registradas en el PDF y en el Excel PRO.")
        st.dataframe(pd.DataFrame([{"Nivel": lv, "Mensaje": ms} for lv, ms in issues]), use_container_width=True)
    else:
        st.success("Sin incoherencias relevantes en este instante.")

    # Diagnóstico automático (nuevo)
    st.subheader("🧠 Diagnóstico automático (explicación técnica)")
    st.dataframe(diag_df[["Categoria","Hallazgo","Severidad","Explicacion"]], use_container_width=True)

    # Gráficas web
    st.subheader("Gráficas (últimos 60 registros)")
    tail = df.tail(60).set_index("time")
    st.line_chart(tail[list(SENSOR_META.keys())])

    # Log eventos web
    st.subheader("Log de eventos")
    ev = pd.DataFrame(st.session_state.events).drop_duplicates()
    st.dataframe(ev.tail(30), use_container_width=True)

    # Descargas
    st.subheader("Descargas")

    # CSV internacional
    st.download_button(
        "📥 Descargar CSV (internacional)",
        data=df.to_csv(index=False).encode("utf-8"),
        file_name="autocheck_sesion.csv",
        mime="text/csv"
    )

    # CSV Excel España (para que no salga todo en una columna)
    csv_excel_es = df.to_csv(index=False, sep=";", decimal=",")
    st.download_button(
        "📥 Descargar CSV (Excel España)",
        data=csv_excel_es.encode("utf-8"),
        file_name="autocheck_sesion_excel_es.csv",
        mime="text/csv"
    )

    vehicle = {
        "vin": st.session_state.get("vin",""),
        "matricula": st.session_state.get("matricula",""),
        "marca_modelo": st.session_state.get("marca_modelo",""),
        "km": st.session_state.get("km", 0),
        "motor": st.session_state.get("motor","Gasolina"),
        "operador": st.session_state.get("operador",""),
        "started_at": st.session_state.get("started_at"),
    }

    # Excel PRO
    xlsx_bytes = make_excel_pro(
        df=df,
        vehicle=vehicle,
        score=score,
        warn_total=warn_total,
        alert_total=alert_total,
        dtc_total=dtc_n,
        diag_rows=diag_rows
    )
    st.download_button(
        "📊 Descargar Excel (.xlsx) PRO",
        data=xlsx_bytes,
        file_name="AutoCheck_Sesion_PRO.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # PDF PRO + diagnóstico automático dentro
    pdf_bytes = build_pdf_report_pro(
        df=df,
        last=last,
        score=score,
        warn_total=warn_total,
        alert_total=alert_total,
        dtc_total=dtc_n,
        recs=recs,
        vehicle=vehicle,
        events_df=ev,
        diag_rows=diag_rows
    )

    st.download_button(
        "🧾 Descargar informe PDF (PRO)",
        data=pdf_bytes,
        file_name="AutoCheck_Informe_Diagnosis_PRO.pdf",
        mime="application/pdf"
    )

# Auto-refresco
if auto:
    time.sleep(1)
    st.rerun()